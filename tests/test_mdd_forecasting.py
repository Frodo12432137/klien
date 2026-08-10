from __future__ import annotations

import tempfile
import unittest
from types import SimpleNamespace
from unittest.mock import patch
from pathlib import Path

import numpy as np
import pandas as pd

from mdd_forecasting.cli import (
    _infer_sql_date_range,
    _parser,
    _prediction_columns,
    _validate_args,
)
from mdd_forecasting.config import PipelineConfig, WEATHER_FEATURES
from mdd_forecasting.database import (
    SqlServerSettings,
    query_weather_sql,
    weather_query_parameters,
)
from mdd_forecasting.generate_demo_data import generate
from mdd_forecasting.features import FeatureSpec, add_lag_features
from mdd_forecasting.io import (
    join_energy_weather,
    normalize_energy,
    normalize_weather,
    prepare_joined_dataset,
    read_energy_file,
)
from mdd_forecasting.model import (
    HybridModelArtifact,
    _baseline_for_fold,
    _blend_prediction,
    _CatBoostTimeLimit,
    _fit_residual_model,
    _select_blend_alpha,
    _split_calibration_window,
    run_forecasting,
)
from mdd_forecasting.report import write_results_workbook
from uruchom_model import (
    BLEND_GRID_STEPS,
    CALIBRATION_DAYS,
    CATBOOST_DEPTH,
    EXECUTION_PROFILE,
    FOLDS,
    INPUT_ROW_SELECTION,
    MAX_FIT_MINUTES,
    MAX_INPUT_ROWS,
    MAX_ITER,
    MAX_TRAIN_ROWS,
    MIN_BLEND_IMPROVEMENT,
    MIN_CALIBRATION_ROWS,
    MODEL_BACKEND,
    MODEL_PROGRESS_INTERVAL,
    MIN_LEAD_HOURS,
    SQL_QUERY_TIMEOUT_SECONDS,
    SQL_CONNECT_TIMEOUT_SECONDS,
    VALIDATION_DAYS,
    WEATHER_AVAILABLE_FROM,
    build_cli_args,
)


class TestMddForecasting(unittest.TestCase):
    def test_click_launcher_builds_portable_arguments(self):
        with tempfile.TemporaryDirectory() as tmp:
            args = build_cli_args(
                Path(tmp) / "energia.xlsx",
                Path(tmp) / "wyniki",
            )
        self.assertEqual(args[args.index("--model-backend") + 1], MODEL_BACKEND)
        self.assertEqual(
            args[args.index("--min-lead-hours") + 1], str(MIN_LEAD_HOURS)
        )
        self.assertEqual(
            args[args.index("--weather-available-from") + 1],
            WEATHER_AVAILABLE_FROM,
        )
        expected_options = {
            "--execution-profile": EXECUTION_PROFILE,
            "--max-input-rows": str(MAX_INPUT_ROWS),
            "--input-row-selection": INPUT_ROW_SELECTION,
            "--validation-days": str(VALIDATION_DAYS),
            "--folds": str(FOLDS),
            "--calibration-days": str(CALIBRATION_DAYS),
            "--min-calibration-rows": str(MIN_CALIBRATION_ROWS),
            "--min-blend-improvement": str(MIN_BLEND_IMPROVEMENT),
            "--blend-grid-steps": str(BLEND_GRID_STEPS),
            "--max-train-rows": str(MAX_TRAIN_ROWS),
            "--max-iter": str(MAX_ITER),
            "--catboost-depth": str(CATBOOST_DEPTH),
            "--max-fit-minutes": str(MAX_FIT_MINUTES),
            "--model-progress-interval": str(MODEL_PROGRESS_INTERVAL),
            "--sql-query-timeout": str(SQL_QUERY_TIMEOUT_SECONDS),
            "--sql-connect-timeout": str(SQL_CONNECT_TIMEOUT_SECONDS),
        }
        for option, expected in expected_options.items():
            self.assertEqual(args[args.index(option) + 1], expected)
        self.assertIn("--skip-importance", args)
        self.assertIn("--compact-output", args)
        sql_path = Path(args[args.index("--weather-sql") + 1])
        self.assertTrue(sql_path.exists())
        self.assertNotIn("10200871", " ".join(args))

    def test_fast_profile_rejects_invalid_limits(self):
        args = _parser().parse_args(
            [
                "--energy",
                "energia.xlsx",
                "--weather",
                "pogoda.csv",
                "--output-dir",
                "wyniki",
            ]
        )
        args.folds = 0
        with self.assertRaisesRegex(ValueError, "dodatnie"):
            _validate_args(args)

        args.folds = 1
        args.min_train_rows = 500
        args.max_train_rows = 100
        with self.assertRaisesRegex(ValueError, "max-train-rows"):
            _validate_args(args)

        args.min_train_rows = 100
        args.max_train_rows = 500
        for invalid_improvement in (-0.01, 1.0):
            args.min_blend_improvement = invalid_improvement
            with self.assertRaisesRegex(ValueError, "min-blend-improvement"):
                _validate_args(args)

        args.min_blend_improvement = 0.02
        args.blend_grid_steps = 1
        with self.assertRaisesRegex(ValueError, "blend-grid-steps"):
            _validate_args(args)

    def test_tail_input_keeps_newest_excel_rows(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "energia.xlsx"
            workbook = Workbook()
            old = workbook.active
            old.title = "Dane_01"
            new = workbook.create_sheet("Dane_02")
            header = ["Nazwa", "Kierunek", "Grupa", "Nazwa", "Kierunek", "Doba", "Rodzaj", "Godzina", "Wartość"]
            for sheet in (old, new):
                sheet.append(header)
            for day in range(1, 4):
                old.append(["BIA", 1, "G", "K", "czynne pobranie", f"2024-01-0{day}", 4, 1, day])
            for day in range(1, 5):
                new.append(["LUB", -1, "G", "K", "czynne oddanie", f"2025-02-0{day}", 4, 1, day])
            workbook.save(path)

            selected = read_energy_file(path, max_rows=3, row_selection="tail")
            self.assertEqual(len(selected), 3)
            self.assertEqual(set(selected["source_sheet"]), {"Dane_02"})
            self.assertEqual(
                selected["doba_handlowa"].astype(str).tolist(),
                ["2025-02-02", "2025-02-03", "2025-02-04"],
            )
            self.assertEqual(
                int(selected.attrs["source_rows_total_estimate"]), 7
            )

    def test_lag_window_uses_same_hour_from_day_3_to_day_14(self):
        dates = pd.date_range("2025-01-01", periods=15, freq="D")
        frame = pd.DataFrame(
            {
                "oddzial_code": "BIA",
                "grupa": "G",
                "klient_nazwa": "KLIENT",
                "kierunek_energii_norm": "POBRANIE",
                "rodzaj": "4",
                "doba_handlowa": dates,
                "godzina_handlowa": 8,
                "wartosc_rzeczywista": range(1, 16),
            }
        )
        with_lags = add_lag_features(frame)
        last = with_lags.iloc[-1]
        for days in range(3, 15):
            self.assertEqual(last[f"lag_{24 * days}h"], 15 - days)
        self.assertNotIn("lag_24h", with_lags.columns)
        self.assertNotIn("lag_48h", with_lags.columns)
        self.assertEqual(last["lag_srednia_3_14_dni"], 6.5)
        self.assertEqual(last["wartosc_bazowa"], 6.5)
        self.assertEqual(last["liczba_lagow_bazowych"], 12.0)
        self.assertAlmostEqual(last["lag_odchylenie_3_14_dni"], 3.6055512755)
        self.assertEqual(last["lag_trend_krotki_vs_dlugi"], 7.5)
        self.assertEqual(last["lag_7d_vs_srednia"], 1.5)

    def test_calibration_window_is_strictly_before_validation(self):
        validation_start = pd.Timestamp("2025-01-10 00:00:00", tz="UTC")
        history = pd.DataFrame(
            {
                "model_timestamp_utc": pd.to_datetime(
                    [
                        "2025-01-07 23:59:59+00:00",
                        "2025-01-08 00:00:00+00:00",
                        "2025-01-09 23:59:59+00:00",
                        "2025-01-10 00:00:00+00:00",
                        None,
                    ],
                    utc=True,
                )
            }
        )
        config = PipelineConfig(calibration_days=2)

        fit, calibration = _split_calibration_window(
            history, validation_start, config
        )

        self.assertEqual(fit.index.tolist(), [0])
        self.assertEqual(calibration.index.tolist(), [1, 2])
        self.assertLess(
            fit["model_timestamp_utc"].max(),
            calibration["model_timestamp_utc"].min(),
        )
        self.assertTrue(
            calibration["model_timestamp_utc"].lt(validation_start).all()
        )

    def test_baseline_preserves_zero_and_uses_hierarchical_train_fallback(self):
        train = pd.DataFrame(
            {
                "oddzial_code": ["BIA", "BIA", "BIA", "LUB", "LUB"],
                "klient_nazwa": ["K1", "K1", "K2", "K3", "K4"],
                "kierunek_energii_norm": [
                    "POBRANIE",
                    "POBRANIE",
                    "POBRANIE",
                    "POBRANIE",
                    "ODDANIE",
                ],
                "godzina": [8, 8, 8, 8, 9],
                "wartosc_rzeczywista": [10.0, 20.0, 30.0, 50.0, 100.0],
            }
        )
        test = pd.DataFrame(
            {
                "wartosc_bazowa": [0.0, np.nan, np.nan, np.nan, np.nan, np.nan],
                "oddzial_code": ["BIA", "BIA", "BIA", "XYZ", "XYZ", "XYZ"],
                "klient_nazwa": ["NOWY", "K1", "NOWY", "NOWY", "NOWY", "NOWY"],
                "kierunek_energii_norm": [
                    "POBRANIE",
                    "POBRANIE",
                    "POBRANIE",
                    "POBRANIE",
                    "ODDANIE",
                    "BRAK",
                ],
                "godzina": [8, 8, 8, 8, 7, 7],
            }
        )

        baseline = _baseline_for_fold(train, test)

        np.testing.assert_array_equal(
            baseline.to_numpy(dtype=float),
            np.array([0.0, 15.0, 20.0, 25.0, 100.0, 30.0]),
        )

    def test_blend_alpha_finds_known_optimum(self):
        selection = _select_blend_alpha(
            actual=np.array([12.0, 8.0, 12.0, 8.0]),
            baseline=np.array([10.0, 10.0, 10.0, 10.0]),
            correction=np.array([4.0, -4.0, 4.0, -4.0]),
            min_rows=1,
            min_improvement=0.0,
            grid_steps=5,
        )

        self.assertEqual(selection.alpha, 0.5)
        self.assertEqual(selection.n, 4)
        self.assertEqual(selection.hybrid_mae, 0.0)
        self.assertEqual(selection.improvement, 1.0)
        self.assertEqual(selection.reason, "HYBRYDA_WYBRANA")

    def test_blend_falls_back_when_correction_is_worse(self):
        selection = _select_blend_alpha(
            actual=np.array([100.0, 120.0, 80.0]),
            baseline=np.array([101.0, 119.0, 81.0]),
            correction=np.array([50.0, -50.0, 50.0]),
            min_rows=1,
            min_improvement=0.0,
            grid_steps=5,
        )

        self.assertEqual(selection.alpha, 0.0)
        self.assertEqual(selection.hybrid_mae, selection.baseline_mae)
        self.assertEqual(selection.reason, "BRAK_POTWIERDZONEJ_POPRAWY")

    def test_blend_tie_prefers_zero_alpha(self):
        selection = _select_blend_alpha(
            actual=np.array([9.0, 11.0]),
            baseline=np.array([10.0, 10.0]),
            correction=np.array([0.0, 0.0]),
            min_rows=1,
            min_improvement=0.0,
            grid_steps=5,
        )

        self.assertEqual(selection.alpha, 0.0)
        self.assertEqual(selection.hybrid_mae, selection.baseline_mae)
        self.assertEqual(selection.reason, "BRAK_POTWIERDZONEJ_POPRAWY")

    def test_blend_requires_minimum_confirmed_improvement(self):
        selection = _select_blend_alpha(
            actual=np.array([10.0, 10.0]),
            baseline=np.array([9.0, 11.0]),
            correction=np.array([0.1, -0.1]),
            min_rows=1,
            min_improvement=0.20,
            grid_steps=11,
        )

        self.assertEqual(selection.alpha, 0.0)
        self.assertAlmostEqual(selection.improvement, 0.10)
        self.assertEqual(selection.reason, "BRAK_POTWIERDZONEJ_POPRAWY")

    def test_residual_model_keeps_signed_target_without_log_transform(self):
        class CapturingEstimator:
            def fit(self, features, target):
                self.features = features.copy()
                self.target = pd.Series(target).copy()
                return self

            def predict(self, features):
                return np.zeros(len(features), dtype=float)

        estimator = CapturingEstimator()
        train = pd.DataFrame(
            {
                "feature": [1.0, 2.0, 3.0, 4.0],
                "wartosc_rzeczywista": [8.0, 25.0, 27.0, 99.0],
                "wartosc_bazowa": [10.0, 20.0, 30.0, np.nan],
            }
        )
        spec = FeatureSpec(categorical=[], numeric=["feature"])
        config = PipelineConfig(min_train_rows=1)

        with patch("mdd_forecasting.model._make_model", return_value=estimator):
            fitted = _fit_residual_model(train, spec, config)

        self.assertIs(fitted, estimator)
        np.testing.assert_array_equal(
            estimator.target.to_numpy(dtype=float),
            np.array([-2.0, 5.0, -3.0]),
        )
        self.assertEqual(estimator.features.index.tolist(), [0, 1, 2])

    def test_blend_prediction_keeps_negative_correction_and_clips_final_value(self):
        predictions = _blend_prediction(
            baseline=np.array([10.0, 2.0]),
            correction=np.array([-4.0, -10.0]),
            alpha=0.5,
        )

        np.testing.assert_array_equal(predictions, np.array([8.0, 0.0]))

    def test_compact_output_contains_hybrid_audit_columns(self):
        audit_columns = {
            "korekta_ml_surowa",
            "wartosc_ml_przed_blendem",
            "blend_alpha",
            "strategia_predykcji",
            "kalibracja_n",
            "kalibracja_poprawa_mae",
            "kalibracja_powod",
        }
        frame = pd.DataFrame(columns=["wartosc_przewidywana", *sorted(audit_columns)])

        selected = set(_prediction_columns(frame, compact=True))

        self.assertTrue(audit_columns.issubset(selected), audit_columns - selected)

    def test_catboost_time_callback_stops_after_deadline(self):
        with patch(
            "mdd_forecasting.model.time.monotonic", side_effect=[0.0, 1.0, 4.0]
        ):
            callback = _CatBoostTimeLimit(3.0)
            self.assertTrue(callback.after_iteration(None))
            self.assertFalse(callback.after_iteration(None))
            self.assertTrue(callback.stopped_by_time)

    def test_fast_profile_runs_at_most_four_bounded_fits(self):
        class FakeModel:
            _mdd_stopped_by_time = False

            def predict(self, frame):
                return [0.0] * len(frame)

        fit_sizes: list[int] = []

        def fake_fit(train, _spec, _config):
            fit_sizes.append(len(train))
            return FakeModel()

        with tempfile.TemporaryDirectory() as tmp:
            energy_path, weather_path = generate(Path(tmp), days=20, seed=9)
            config = PipelineConfig(
                execution_profile="fast_30min",
                validation_days=3,
                n_splits=1,
                min_train_rows=50,
                max_train_rows=100,
                max_iter=60,
                compute_importance=False,
            )
            joined, _, _ = prepare_joined_dataset(
                energy_path, weather_path, config=config
            )
            messages: list[str] = []
            with patch("mdd_forecasting.model._fit_residual_model", side_effect=fake_fit):
                run_forecasting(joined, config, progress_callback=messages.append)

        self.assertEqual(len(fit_sizes), 4)
        self.assertLessEqual(max(fit_sizes), 100)
        self.assertTrue(any("12 lagów" in message for message in messages))
        self.assertTrue(any("Model końcowy POBRANIE" in message for message in messages))

    def test_outer_oof_target_does_not_change_fold_alpha(self):
        with tempfile.TemporaryDirectory() as tmp:
            energy_path, weather_path = generate(Path(tmp), days=10, seed=19)
            config = PipelineConfig(
                validation_days=2,
                n_splits=1,
                calibration_days=2,
                min_calibration_rows=1,
                min_blend_improvement=0.0,
                blend_grid_steps=5,
                min_train_rows=50,
                max_train_rows=50_000,
                max_iter=5,
                compute_importance=False,
            )
            joined, _, _ = prepare_joined_dataset(
                energy_path, weather_path, config=config
            )
            lagged = add_lag_features(joined, lag_days=config.lag_days)
            correction_by_index = (
                lagged["wartosc_rzeczywista"] - lagged["wartosc_bazowa"]
            )

            class IndexedCorrectionModel:
                _mdd_stopped_by_time = False

                def predict(self, features):
                    return (
                        correction_by_index.reindex(features.index)
                        .fillna(0.0)
                        .to_numpy(dtype=float)
                    )

            def fake_fit(_train, _spec, _config):
                return IndexedCorrectionModel()

            actual_times = pd.to_datetime(
                joined["model_timestamp_utc"], errors="coerce", utc=True
            )
            validation_end = actual_times.max().floor("D") + pd.Timedelta(days=1)
            validation_start = validation_end - pd.Timedelta(days=2)
            outer_mask = actual_times.ge(validation_start) & actual_times.lt(
                validation_end
            )
            modified = joined.copy()
            modified.loc[outer_mask, "wartosc_rzeczywista"] = lagged.loc[
                outer_mask, "wartosc_bazowa"
            ].to_numpy()

            with patch(
                "mdd_forecasting.model._fit_residual_model", side_effect=fake_fit
            ):
                original_result = run_forecasting(joined, config)
                modified_result = run_forecasting(modified, config)

        def fold_alphas(result):
            oof = result.predictions[
                result.predictions["status_predykcji"].eq("OOF_BACKTEST")
            ]
            return (
                oof.groupby(["fold", "kierunek_energii_norm"])["blend_alpha"]
                .first()
                .sort_index()
            )

        original_alphas = fold_alphas(original_result)
        modified_alphas = fold_alphas(modified_result)
        self.assertTrue(original_alphas.gt(0.0).all())
        pd.testing.assert_series_equal(original_alphas, modified_alphas)

    def test_sql_parameters_and_integrated_connection(self):
        params = weather_query_parameters("2025-01-01", "2025-02-01", 24)
        self.assertEqual(len(params), 5)
        self.assertEqual(params[2], 24)
        self.assertEqual(params[3:], ["PGESA", "Open Meteo"])
        connection = SqlServerSettings().connection_string()
        self.assertIn("SERVER=MISDWHPRD.GKPGE.PL", connection)
        self.assertIn("DATABASE=PGESA_MarketAnalytics", connection)
        self.assertIn("Trusted_Connection=yes", connection)
        self.assertNotIn("PWD=", connection.upper())
        sql_text = (
            Path("mdd_forecasting/sql/pogoda_mdd.sql").read_text(encoding="utf-8")
        )
        self.assertEqual(sql_text.count("?"), 5)
        self.assertIn("[PGESA_MarketAnalytics].[wa].[vPogodaPrognoza]", sql_text)
        self.assertIn("czasDanychZrodlaUTC", sql_text)

    def test_sql_execution_is_parameterized_and_connection_is_closed(self):
        class FakeConnection:
            def __init__(self):
                self.closed = False
                self.timeout = 0

            def close(self):
                self.closed = True

        connection = FakeConnection()
        fake_pyodbc = SimpleNamespace(
            drivers=lambda: ["ODBC Driver 17 for SQL Server"],
            connect=lambda *args, **kwargs: connection,
        )
        expected = pd.DataFrame({"punkt": ["Lublin"]})
        with patch.dict("sys.modules", {"pyodbc": fake_pyodbc}), patch(
            "mdd_forecasting.database.pd.read_sql_query", return_value=expected
        ) as read_query:
            result = query_weather_sql(
                "mdd_forecasting/sql/pogoda_mdd.sql",
                settings=SqlServerSettings(),
                valid_from_cet="2025-01-01",
                valid_to_cet_exclusive="2025-01-03",
                min_lead_hours=24,
                query_timeout_seconds=120,
            )
        self.assertEqual(result.to_dict("records"), [{"punkt": "Lublin"}])
        self.assertTrue(connection.closed)
        self.assertEqual(connection.timeout, 120)
        params = read_query.call_args.kwargs["params"]
        self.assertEqual(len(params), 5)
        self.assertEqual(params[2:], [24, "PGESA", "Open Meteo"])

    def test_plain_sql_without_placeholders_is_supported(self):
        class FakeConnection:
            def __init__(self):
                self.closed = False

            def close(self):
                self.closed = True

        connection = FakeConnection()
        fake_pyodbc = SimpleNamespace(
            drivers=lambda: ["ODBC Driver 17 for SQL Server"],
            connect=lambda *args, **kwargs: connection,
        )
        expected = pd.DataFrame({"punkt": ["Łódź"]})
        with tempfile.TemporaryDirectory() as tmp:
            sql_path = Path(tmp) / "firmowy.sql"
            sql_path.write_text(
                "DECLARE @data_start date = '2025-01-01'; SELECT 'Łódź' AS punkt;",
                encoding="utf-8",
            )
            with patch.dict("sys.modules", {"pyodbc": fake_pyodbc}), patch(
                "mdd_forecasting.database.pd.read_sql_query", return_value=expected
            ) as read_query, self.assertWarnsRegex(UserWarning, "bez zmian"):
                result = query_weather_sql(
                    sql_path,
                    settings=SqlServerSettings(),
                    valid_from_cet="2025-01-01",
                    valid_to_cet_exclusive="2025-01-03",
                    min_lead_hours=24,
                )
        self.assertEqual(result.to_dict("records"), [{"punkt": "Łódź"}])
        self.assertTrue(connection.closed)
        self.assertNotIn("params", read_query.call_args.kwargs)

    def test_sql_range_is_inferred_from_energy_dates(self):
        energy = pd.DataFrame(
            {"doba_handlowa": ["2025-01-05", "2025-01-07", None]}
        )
        valid_from, valid_to = _infer_sql_date_range(energy, None, None)
        self.assertEqual(valid_from, pd.Timestamp("2025-01-05"))
        self.assertEqual(valid_to, pd.Timestamp("2025-01-08"))

    def test_sql_range_is_clamped_to_weather_start(self):
        energy = pd.DataFrame(
            {"doba_handlowa": ["2023-07-29", "2024-11-15", None]}
        )
        valid_from, valid_to = _infer_sql_date_range(
            energy, None, None, "2024-10-01"
        )
        self.assertEqual(valid_from, pd.Timestamp("2024-10-01"))
        self.assertEqual(valid_to, pd.Timestamp("2024-11-16"))

    def test_sql_is_skipped_when_all_energy_is_before_weather_start(self):
        energy = pd.DataFrame(
            {"doba_handlowa": ["2023-07-29", "2024-09-30", None]}
        )
        valid_from, valid_to = _infer_sql_date_range(
            energy, None, None, "2024-10-01"
        )
        self.assertIsNone(valid_from)
        self.assertEqual(valid_to, pd.Timestamp("2024-10-01"))

    def test_weather_vintage_respects_minimum_lead(self):
        valid = pd.Timestamp("2025-01-03 12:00:00")
        weather = pd.DataFrame(
            {
                "punkt": ["Lublin", "Lublin"],
                "dataGodzinaCET": [valid, valid],
                "czasDanychZrodlaCET": [
                    valid - pd.Timedelta(hours=30),
                    valid - pd.Timedelta(hours=12),
                ],
                "temperatura": [1.0, 99.0],
            }
        )
        selected = normalize_weather(weather, PipelineConfig(min_lead_hours=24))
        self.assertEqual(len(selected), 1)
        self.assertEqual(float(selected.iloc[0]["temperatura"]), 1.0)
        self.assertEqual(float(selected.iloc[0]["weather_lead_hours"]), 30.0)

    def test_weather_before_october_is_expected_and_nulls_are_preserved(self):
        energy_raw = pd.DataFrame(
            [
                ["Dane_01", 2, "BIA", 1, "G", "K", "czynne pobranie", "2024-09-30", 4, 1, 10],
                ["Dane_01", 3, "BIA", 1, "G", "K", "czynne pobranie", "2024-10-02", 4, 1, 11],
                ["Dane_01", 4, "XYZ", 1, "G", "K", "czynne pobranie", "2024-09-30", 4, 1, 12],
            ],
            columns=[
                "source_sheet", "source_row", "oddzial_code", "kierunek_code", "grupa",
                "klient_nazwa", "kierunek_energii", "doba_handlowa", "rodzaj",
                "godzina_handlowa", "wartosc_rzeczywista",
            ],
        )
        energy, _ = normalize_energy(energy_raw, {"bia": "Białystok"})
        valid = pd.Timestamp("2024-10-02 00:00:00")
        weather_raw = pd.DataFrame(
            {
                "punkt": ["Białystok"],
                "dataGodzinaCET": [valid],
                "czasDanychZrodlaCET": [valid - pd.Timedelta(hours=30)],
                "temperatura": [None],
            }
        )
        weather = normalize_weather(weather_raw, PipelineConfig(min_lead_hours=24))
        joined, quality = join_energy_weather(
            energy, weather, weather_available_from="2024-10-01"
        )

        self.assertEqual(len(joined), 3)
        self.assertEqual(joined.loc[0, "weather_status"], "PRZED_STARTEM_POGODY")
        self.assertEqual(joined.loc[1, "weather_status"], "DOPASOWANA")
        self.assertEqual(joined.loc[2, "weather_status"], "BRAK_MAPOWANIA")
        self.assertEqual(float(joined.loc[1, "pogoda_dostepna"]), 1.0)
        self.assertTrue(pd.isna(joined.loc[1, "temperatura"]))
        self.assertEqual(int(joined.loc[1, "liczba_cech_pogodowych"]), 0)
        quality_counts = quality.set_index("kontrola")["liczba"]
        self.assertEqual(int(quality_counts["pogoda_przed_startem_zrodla"]), 1)
        self.assertEqual(int(quality_counts["pogoda_rekord_bez_cech"]), 1)

    def test_end_to_end_demo_creates_oof_predictions(self):
        with tempfile.TemporaryDirectory() as tmp:
            energy_path, weather_path = generate(Path(tmp), days=20, seed=7)
            config = PipelineConfig(
                min_lead_hours=24,
                validation_days=3,
                n_splits=1,
                min_train_rows=100,
                max_train_rows=50_000,
                max_iter=35,
                compute_importance=False,
            )
            joined, quality, _ = prepare_joined_dataset(
                energy_path, weather_path, config=config
            )
            self.assertEqual(len(joined), 20 * 24 * 4 * 2 * 2)
            self.assertTrue(joined["pogoda_dopasowana"].all())
            historical_gap_index = joined.index[100]
            joined.loc[historical_gap_index, "wartosc_rzeczywista"] = float("nan")
            future_time = joined["model_timestamp_utc"].max()
            future_rows = joined["model_timestamp_utc"].eq(future_time)
            joined.loc[future_rows, "wartosc_rzeczywista"] = float("nan")
            joined.loc[future_rows, "pogoda_dopasowana"] = False
            joined.loc[future_rows, "pogoda_dostepna"] = 0.0
            joined.loc[future_rows, "weather_status"] = "BRAK_POGODY_W_ZAKRESIE"
            joined.loc[future_rows, WEATHER_FEATURES] = float("nan")
            result = run_forecasting(joined, config)
            evaluated = result.predictions["status_predykcji"].eq("OOF_BACKTEST")
            self.assertGreater(int(evaluated.sum()), 0)
            self.assertFalse(result.metrics.empty)
            self.assertIn("POGODA", set(result.metrics["zakres"]))
            self.assertIn("HYBRYDA_OOF", set(result.metrics["model"]))
            self.assertIn(
                "SREDNIA_ANALOGICZNYCH_D3_D14", set(result.metrics["model"])
            )
            global_metrics = result.metrics[
                result.metrics["zakres"].eq("GLOBAL")
                & result.metrics["wartosc_zakresu"].eq("ALL")
            ].set_index("model")
            self.assertEqual(
                int(global_metrics.loc["HYBRYDA_OOF", "n"]),
                int(global_metrics.loc["SREDNIA_ANALOGICZNYCH_D3_D14", "n"]),
            )
            self.assertEqual(set(result.models), {"POBRANIE", "ODDANIE"})
            self.assertTrue(
                all(
                    isinstance(artifact, HybridModelArtifact)
                    for artifact in result.models.values()
                )
            )
            self.assertTrue(
                all(0.0 <= artifact.alpha <= 1.0 for artifact in result.models.values())
            )
            expected_audit_columns = {
                "korekta_ml_surowa",
                "wartosc_ml_przed_blendem",
                "blend_alpha",
                "strategia_predykcji",
                "kalibracja_n",
                "kalibracja_poprawa_mae",
                "kalibracja_powod",
            }
            self.assertTrue(
                expected_audit_columns.issubset(result.predictions.columns),
                expected_audit_columns - set(result.predictions.columns),
            )
            expected_lag_features = {
                "liczba_lagow_bazowych",
                "lag_odchylenie_3_14_dni",
                "lag_trend_krotki_vs_dlugi",
                "lag_7d_vs_srednia",
            }
            self.assertTrue(
                expected_lag_features.issubset(result.feature_spec.numeric),
                expected_lag_features - set(result.feature_spec.numeric),
            )
            future_predictions = result.predictions.loc[future_rows]
            self.assertTrue(
                future_predictions["status_predykcji"].eq("PROGNOZA_PRZYSZLA").all()
            )
            self.assertTrue(future_predictions["prognoza_bez_pogody"].all())
            self.assertTrue(future_predictions["blend_alpha"].between(0.0, 1.0).all())
            self.assertTrue(future_predictions["strategia_predykcji"].notna().all())
            self.assertEqual(
                result.predictions.loc[historical_gap_index, "status_predykcji"],
                "HISTORYCZNY_BRAK_TARGETU",
            )
            workbook_path = write_results_workbook(
                predictions=result.predictions,
                prediction_columns=[
                    "source_sheet",
                    "source_row",
                    "klient_nazwa",
                    "wartosc_rzeczywista",
                    "wartosc_przewidywana",
                    "status_predykcji",
                ],
                metrics=result.metrics,
                feature_importance=result.feature_importance,
                quality=quality,
                mapping=pd.DataFrame(
                    {"oddzial_code": ["BIA"], "punkt": ["Białystok"], "status": ["active"]}
                ),
                models=result.models,
                manifest={"test": True},
                output_path=Path(tmp) / "wyniki_mdd.xlsx",
            )
            self.assertTrue(workbook_path.exists())
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            try:
                self.assertIn("Podsumowanie", workbook.sheetnames)
                self.assertIn("Metryki", workbook.sheetnames)
                self.assertTrue(any(name.startswith("Pred_") for name in workbook.sheetnames))
                summary = workbook["Podsumowanie"]
                self.assertEqual(summary["A1"].value, "sekcja")
                first_row = next(summary.iter_rows(values_only=True))
                self.assertEqual(len(first_row), 4)
            finally:
                workbook.close()

    def test_dst_repeated_hour_uses_trade_hour_and_utc(self):
        energy_raw = pd.DataFrame(
            [
                ["Dane_01", 2, "BIA", 1, "G", "K", "czynne pobranie", "2025-10-26", 4, 3, 10],
                ["Dane_01", 3, "BIA", 1, "G", "K", "czynne pobranie", "2025-10-26", 4, 25, 11],
            ],
            columns=[
                "source_sheet", "source_row", "oddzial_code", "kierunek_code", "grupa",
                "klient_nazwa", "kierunek_energii", "doba_handlowa", "rodzaj",
                "godzina_handlowa", "wartosc_rzeczywista",
            ],
        )
        energy, _ = normalize_energy(energy_raw, {"bia": "Białystok"})
        local_valid = pd.Timestamp("2025-10-26 02:00:00")
        weather_raw = pd.DataFrame(
            {
                "punkt": ["Białystok", "Białystok"],
                "dataCET": ["2025-10-26", "2025-10-26"],
                "godzinaHandlowa25": [3, 25],
                "dataGodzinaCET": [local_valid, local_valid],
                "dataGodzinaUTC": ["2025-10-26T00:00:00Z", "2025-10-26T01:00:00Z"],
                "czasDanychZrodlaCET": [
                    local_valid - pd.Timedelta(hours=30),
                    local_valid - pd.Timedelta(hours=30),
                ],
                "temperatura": [5, 6],
            }
        )
        weather = normalize_weather(weather_raw, PipelineConfig(min_lead_hours=24))
        joined, _ = join_energy_weather(energy, weather)
        self.assertEqual(len(joined), 2)
        self.assertEqual(joined["model_timestamp_utc"].nunique(), 2)
        self.assertTrue(joined["pogoda_dopasowana"].all())


if __name__ == "__main__":
    unittest.main()
