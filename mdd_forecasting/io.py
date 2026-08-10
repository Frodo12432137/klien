from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd

from .config import ENERGY_COLUMNS, PipelineConfig, WEATHER_ALIASES, WEATHER_FEATURES


def _ascii_token(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = str(value).strip().translate(str.maketrans({"Ł": "L", "ł": "l"}))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-zA-Z0-9]+", "", text).lower()


def _clean_text(series: pd.Series) -> pd.Series:
    out = series.astype("string").str.strip().str.replace(r"\s+", " ", regex=True)
    return out.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "<NA>": pd.NA})


def _numeric(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype("string")
        .str.replace("\u00a0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def load_branch_mapping(path: str | Path) -> tuple[dict[str, str], pd.DataFrame]:
    mapping_df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    required = {"oddzial_code", "punkt", "status"}
    missing = required - set(mapping_df.columns)
    if missing:
        raise ValueError(f"Plik mapowania nie ma kolumn: {sorted(missing)}")

    active_mask = mapping_df["status"].astype(str).str.lower().eq("active")
    active = mapping_df.loc[active_mask].copy()
    active["_code_key"] = active["oddzial_code"].map(_ascii_token)
    conflicts = active.groupby("_code_key")["punkt"].nunique(dropna=False)
    conflicts = conflicts[conflicts.gt(1)]
    if not conflicts.empty:
        dup = active.loc[active["_code_key"].isin(conflicts.index), "oddzial_code"].tolist()
        raise ValueError(f"Sprzeczne aktywne mapowania kodów oddziałów: {dup}")
    unique_active = active.drop_duplicates("_code_key", keep="first")
    return dict(zip(unique_active["_code_key"], unique_active["punkt"])), mapping_df


def _energy_frame_from_values(
    values: Iterable[Iterable[object]], source_sheet: str, start_row: int = 2
) -> pd.DataFrame:
    rows: list[list[object]] = []
    source_rows: list[int] = []
    for row_number, row in enumerate(values, start=start_row):
        selected = list(row[: len(ENERGY_COLUMNS)])
        if len(selected) < len(ENERGY_COLUMNS):
            selected.extend([None] * (len(ENERGY_COLUMNS) - len(selected)))
        if not any(value is not None and str(value).strip() != "" for value in selected):
            continue
        rows.append(selected)
        source_rows.append(row_number)
    frame = pd.DataFrame(rows, columns=ENERGY_COLUMNS)
    frame.insert(0, "source_row", source_rows)
    frame.insert(0, "source_sheet", source_sheet)
    return frame


def read_energy_file(
    path: str | Path,
    sheet_regex: str = r"^Dane_\d+$",
    max_rows: int | None = None,
) -> pd.DataFrame:
    """Czyta A:I pozycyjnie, ponieważ w źródle dwa nagłówki mają nazwę `Nazwa`."""

    source = Path(path)
    suffix = source.suffix.lower()
    if suffix in {".csv", ".txt"}:
        raw = pd.read_csv(source, sep=None, engine="python", encoding="utf-8-sig")
        if set(ENERGY_COLUMNS).issubset(raw.columns):
            frame = raw[ENERGY_COLUMNS].copy()
        elif raw.shape[1] >= len(ENERGY_COLUMNS):
            frame = raw.iloc[:, : len(ENERGY_COLUMNS)].copy()
            frame.columns = ENERGY_COLUMNS
        else:
            raise ValueError(
                f"Plik energii ma {raw.shape[1]} kolumn, a potrzeba co najmniej 9 (A:I)."
            )
        frame.insert(0, "source_row", np.arange(2, len(frame) + 2))
        frame.insert(0, "source_sheet", source.stem)
        return frame.head(max_rows) if max_rows else frame

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Dane energii muszą być plikiem CSV, XLSX albo XLSM.")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - komunikat dla środowiska użytkownika
        raise RuntimeError("Do strumieniowego odczytu XLSX wymagany jest openpyxl.") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    pattern = re.compile(sheet_regex, flags=re.IGNORECASE)
    selected_sheets = [name for name in workbook.sheetnames if pattern.search(name)]
    if not selected_sheets:
        raise ValueError(
            f"Nie znaleziono arkuszy pasujących do {sheet_regex!r}. Dostępne: {workbook.sheetnames}"
        )

    frames: list[pd.DataFrame] = []
    rows_left = max_rows
    try:
        for sheet_name in selected_sheets:
            sheet = workbook[sheet_name]
            values = sheet.iter_rows(min_row=2, max_col=len(ENERGY_COLUMNS), values_only=True)
            if rows_left is not None:
                limited = []
                for row in values:
                    if len(limited) >= rows_left:
                        break
                    limited.append(row)
                values = limited
            frame = _energy_frame_from_values(values, sheet_name)
            frames.append(frame)
            if rows_left is not None:
                rows_left -= len(frame)
                if rows_left <= 0:
                    break
    finally:
        workbook.close()

    if not frames:
        return pd.DataFrame(columns=["source_sheet", "source_row", *ENERGY_COLUMNS])
    return pd.concat(frames, ignore_index=True)


def normalize_energy(
    energy: pd.DataFrame, branch_mapping: dict[str, str]
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = energy.copy()
    for col in ["oddzial_code", "grupa", "klient_nazwa", "kierunek_energii", "rodzaj"]:
        df[col] = _clean_text(df[col])

    df["wartosc_rzeczywista"] = _numeric(df["wartosc_rzeczywista"])
    df["kierunek_code"] = _numeric(df["kierunek_code"])
    df["godzina_handlowa"] = _numeric(
        df["godzina_handlowa"].astype("string").str.extract(r"(\d+)", expand=False)
    )
    df["doba_handlowa"] = pd.to_datetime(df["doba_handlowa"], errors="coerce").dt.normalize()

    direction_text = df["kierunek_energii"].map(_ascii_token)
    df["kierunek_energii_norm"] = pd.Series(pd.NA, index=df.index, dtype="string")
    df.loc[direction_text.str.contains("pobran", na=False), "kierunek_energii_norm"] = "POBRANIE"
    df.loc[direction_text.str.contains("oddani", na=False), "kierunek_energii_norm"] = "ODDANIE"
    df["direction_source"] = np.where(df["kierunek_energii_norm"].notna(), "kolumna_E", "brak")

    # Awaryjny fallback służy wyłącznie brakującym E; E pozostaje źródłem prawdy.
    missing_direction = df["kierunek_energii_norm"].isna()
    df.loc[missing_direction & df["kierunek_code"].gt(0), "kierunek_energii_norm"] = "POBRANIE"
    df.loc[missing_direction & df["kierunek_code"].lt(0), "kierunek_energii_norm"] = "ODDANIE"
    df.loc[missing_direction & df["kierunek_energii_norm"].notna(), "direction_source"] = (
        "fallback_kolumna_B"
    )

    df["_oddzial_key"] = df["oddzial_code"].map(_ascii_token)
    df["punkt"] = df["_oddzial_key"].map(branch_mapping).astype("string")
    df["_punkt_key"] = df["punkt"].map(_ascii_token)

    normal_hour = df["godzina_handlowa"].between(1, 24, inclusive="both")
    df["valid_timestamp"] = pd.NaT
    df.loc[normal_hour, "valid_timestamp"] = (
        df.loc[normal_hour, "doba_handlowa"]
        + pd.to_timedelta(df.loc[normal_hour, "godzina_handlowa"] - 1, unit="h")
    )

    mismatch = (
        df["kierunek_energii_norm"].eq("POBRANIE") & df["kierunek_code"].lt(0)
    ) | (df["kierunek_energii_norm"].eq("ODDANIE") & df["kierunek_code"].gt(0))

    unknown_codes = sorted(
        df.loc[df["oddzial_code"].notna() & df["punkt"].isna(), "oddzial_code"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    quality = pd.DataFrame(
        [
            ["wiersze_energii", len(df), ""],
            ["brak_kodu_oddzialu", int(df["oddzial_code"].isna().sum()), ""],
            ["nieznany_kod_oddzialu", int((df["oddzial_code"].notna() & df["punkt"].isna()).sum()), ", ".join(unknown_codes)],
            ["brak_kierunku_E_i_B", int(df["kierunek_energii_norm"].isna().sum()), ""],
            ["niezgodnosc_B_wobec_E", int(mismatch.sum()), "kolumna E ma pierwszeństwo"],
            ["godzina_25_lub_nieprawidlowa", int((~normal_hour).sum()), "wymaga klucza godzinaHandlowa25"],
            ["brak_wartosci_rzeczywistej", int(df["wartosc_rzeczywista"].isna().sum()), "może oznaczać wiersze przyszłe"],
            ["ujemna_wartosc", int(df["wartosc_rzeczywista"].lt(0).sum()), "wartości nie są automatycznie odwracane"],
        ],
        columns=["kontrola", "liczba", "szczegoly"],
    )
    return df, quality


def read_weather_file(path: str | Path) -> pd.DataFrame:
    source = Path(path)
    if source.suffix.lower() in {".csv", ".txt"}:
        return pd.read_csv(source, sep=None, engine="python", encoding="utf-8-sig")
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return pd.read_excel(source)
    raise ValueError("Dane pogodowe muszą być plikiem CSV, XLSX albo XLSM.")


def normalize_weather(weather: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    raw = weather.copy()
    normalized_headers = {_ascii_token(col): col for col in raw.columns}
    rename: dict[object, str] = {}
    for canonical, aliases in WEATHER_ALIASES.items():
        for alias in aliases:
            source_col = normalized_headers.get(_ascii_token(alias))
            if source_col is not None:
                rename[source_col] = canonical
                break
    df = raw.rename(columns=rename)

    for required in ["punkt", "weather_valid_time"]:
        if required not in df.columns:
            raise ValueError(
                f"Brak kolumny pogodowej {required!r}. Dostępne nagłówki: {list(raw.columns)}"
            )
    if not config.weather_already_vintaged and "weather_issue_time" not in df.columns:
        raise ValueError(
            "Brak czasu wydania prognozy `czasDanychZrodlaCET`. Bez niego backtest grozi leakage. "
            "Jeżeli eksport został już poprawnie zwintagowany w SQL, użyj --weather-already-vintaged."
        )

    df["punkt"] = _clean_text(df["punkt"])
    df["_punkt_key"] = df["punkt"].map(_ascii_token)
    df["weather_valid_time"] = pd.to_datetime(df["weather_valid_time"], errors="coerce")
    if "weather_valid_time_utc" in df.columns:
        df["weather_valid_time_utc"] = pd.to_datetime(
            df["weather_valid_time_utc"], errors="coerce", utc=True
        )
    else:
        df["weather_valid_time_utc"] = pd.NaT
    if "weather_issue_time" in df.columns:
        df["weather_issue_time"] = pd.to_datetime(df["weather_issue_time"], errors="coerce")
    else:
        df["weather_issue_time"] = pd.NaT
    if "weather_issue_time_utc" in df.columns:
        df["weather_issue_time_utc"] = pd.to_datetime(
            df["weather_issue_time_utc"], errors="coerce", utc=True
        )
    else:
        df["weather_issue_time_utc"] = pd.NaT
    explicit_trade_key = {
        "weather_trade_date",
        "weather_trade_hour",
    }.issubset(df.columns)
    if "weather_trade_date" in df.columns:
        df["weather_trade_date"] = pd.to_datetime(df["weather_trade_date"], errors="coerce").dt.normalize()
    else:
        df["weather_trade_date"] = df["weather_valid_time"].dt.normalize()
    if "weather_trade_hour" in df.columns:
        df["weather_trade_hour"] = _numeric(
            df["weather_trade_hour"].astype("string").str.extract(r"(\d+)", expand=False)
        )
    else:
        df["weather_trade_hour"] = df["weather_valid_time"].dt.hour + 1

    for feature in WEATHER_FEATURES:
        if feature not in df.columns:
            df[feature] = np.nan
        df[feature] = _numeric(df[feature])

    df = df[df["weather_valid_time"].notna() & df["_punkt_key"].ne("")].copy()
    if not config.weather_already_vintaged:
        lead_delta = pd.to_timedelta(config.min_lead_hours, unit="h")
        has_utc = df["weather_valid_time_utc"].notna() & df[
            "weather_issue_time_utc"
        ].notna()
        eligible = pd.Series(False, index=df.index)
        if has_utc.any():
            eligible.loc[has_utc] = df.loc[has_utc, "weather_issue_time_utc"].le(
                df.loc[has_utc, "weather_valid_time_utc"] - lead_delta
            )
        has_local = ~has_utc & df["weather_issue_time"].notna()
        eligible.loc[has_local] = df.loc[has_local, "weather_issue_time"].le(
            df.loc[has_local, "weather_valid_time"] - lead_delta
        )
        df = df[eligible].copy()

    df["_weather_explicit_trade_key"] = explicit_trade_key
    key = (
        ["_punkt_key", "weather_trade_date", "weather_trade_hour"]
        if explicit_trade_key
        else ["_punkt_key", "weather_valid_time"]
    )
    sort_cols = [*key, "weather_issue_time_utc", "weather_issue_time"]
    df = df.sort_values(sort_cols, na_position="first")
    df = df.drop_duplicates(key, keep="last")
    df["weather_lead_hours"] = np.nan
    has_utc = df["weather_valid_time_utc"].notna() & df["weather_issue_time_utc"].notna()
    if has_utc.any():
        df.loc[has_utc, "weather_lead_hours"] = (
            df.loc[has_utc, "weather_valid_time_utc"]
            - df.loc[has_utc, "weather_issue_time_utc"]
        ).dt.total_seconds() / 3600.0
    has_local = ~has_utc & df["weather_issue_time"].notna()
    df.loc[has_local, "weather_lead_hours"] = (
        df.loc[has_local, "weather_valid_time"]
        - df.loc[has_local, "weather_issue_time"]
    ).dt.total_seconds() / 3600.0
    df["_weather_row_id"] = np.arange(len(df), dtype=np.int64)
    return df


def join_energy_weather(
    energy: pd.DataFrame, weather: pd.DataFrame, timezone: str = "Europe/Warsaw"
) -> tuple[pd.DataFrame, pd.DataFrame]:
    before = len(energy)
    use_trade_key = (
        weather["_weather_explicit_trade_key"].any()
        and weather["weather_trade_date"].notna().any()
        and weather["weather_trade_hour"].notna().any()
        and energy["doba_handlowa"].notna().any()
        and energy["godzina_handlowa"].notna().any()
    )
    if use_trade_key:
        right_key = ["_punkt_key", "weather_trade_date", "weather_trade_hour"]
        if weather.duplicated(right_key).any():
            raise ValueError("Pogoda nie jest unikalna po punkt + dataCET + godzinaHandlowa25.")
        joined = energy.merge(
            weather,
            how="left",
            left_on=["_punkt_key", "doba_handlowa", "godzina_handlowa"],
            right_on=right_key,
            suffixes=("", "_weather"),
            validate="many_to_one",
        )
        join_mode = "punkt + dataCET + godzinaHandlowa25"
    else:
        right_key = ["_punkt_key", "weather_valid_time"]
        if weather.duplicated(right_key).any():
            raise ValueError("Pogoda nie jest unikalna po punkt + dataGodzinaCET.")
        joined = energy.merge(
            weather,
            how="left",
            left_on=["_punkt_key", "valid_timestamp"],
            right_on=right_key,
            suffixes=("", "_weather"),
            validate="many_to_one",
        )
        join_mode = "punkt + dataGodzinaCET"

    if len(joined) != before:
        raise AssertionError("Łączenie pogody zmieniło liczbę wierszy energii.")
    # Dla 25. godziny uproszczone doba + godzina-1 jest niejednoznaczne. Po
    # bezpiecznym joinie po godzinie handlowej przejmujemy valid time z pogody.
    missing_valid = joined["valid_timestamp"].isna() & joined["weather_valid_time"].notna()
    joined.loc[missing_valid, "valid_timestamp"] = joined.loc[missing_valid, "weather_valid_time"]
    joined["model_timestamp_utc"] = pd.to_datetime(
        joined["weather_valid_time_utc"], errors="coerce", utc=True
    )
    missing_utc = joined["model_timestamp_utc"].isna() & joined["valid_timestamp"].notna()
    if missing_utc.any():
        localized = (
            pd.to_datetime(joined.loc[missing_utc, "valid_timestamp"], errors="coerce")
            .dt.tz_localize(timezone, ambiguous="NaT", nonexistent="NaT")
            .dt.tz_convert("UTC")
        )
        joined.loc[missing_utc, "model_timestamp_utc"] = localized
    joined["pogoda_dopasowana"] = joined["_weather_row_id"].notna()
    quality = pd.DataFrame(
        [
            ["tryb_laczenia", before, join_mode],
            ["pogoda_dopasowana", int(joined["pogoda_dopasowana"].sum()), ""],
            ["pogoda_niedopasowana", int((~joined["pogoda_dopasowana"]).sum()), ""],
        ],
        columns=["kontrola", "liczba", "szczegoly"],
    )
    return joined, quality


def prepare_joined_dataset(
    energy_path: str | Path,
    weather_path: str | Path,
    config: PipelineConfig,
    max_rows: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    energy, energy_quality, mapping_df = prepare_energy_dataset(
        energy_path, config=config, max_rows=max_rows
    )
    return join_prepared_energy_with_weather(
        energy,
        energy_quality,
        mapping_df,
        read_weather_file(weather_path),
        config,
    )


def prepare_energy_dataset(
    energy_path: str | Path,
    config: PipelineConfig,
    max_rows: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Czyta i normalizuje energię przed zapytaniem SQL zależnym od zakresu dat."""

    branch_mapping, mapping_df = load_branch_mapping(config.mapping_path)
    energy_raw = read_energy_file(energy_path, max_rows=max_rows)
    energy, energy_quality = normalize_energy(energy_raw, branch_mapping)
    return energy, energy_quality, mapping_df


def join_prepared_energy_with_weather(
    energy: pd.DataFrame,
    energy_quality: pd.DataFrame,
    mapping_df: pd.DataFrame,
    weather_raw: pd.DataFrame,
    config: PipelineConfig,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Łączy wcześniej przygotowaną energię z pogodą pobraną z pliku albo SQL."""

    weather = normalize_weather(weather_raw, config)
    joined, join_quality = join_energy_weather(energy, weather, timezone=config.timezone)
    quality = pd.concat([energy_quality, join_quality], ignore_index=True)
    return joined, quality, mapping_df
