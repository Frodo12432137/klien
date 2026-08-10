from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd

from .config import WEATHER_FEATURES


EXCEL_MAX_ROWS = 1_048_576
EXCEL_DATA_ROWS_PER_SHEET = EXCEL_MAX_ROWS - 1


def _excel_scalar(value: object) -> object:
    """Konwertuje typy pandas/numpy na wartości obsługiwane przez Excel."""

    if value is None or value is pd.NA:
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        if value.tzinfo is not None:
            value = value.tz_convert("UTC").tz_localize(None)
        return value.to_pydatetime()
    if isinstance(value, np.datetime64):
        return _excel_scalar(pd.Timestamp(value))
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _safe_sheet_name(raw_name: object, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", str(raw_name)).strip(" '") or "Dane"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate.casefold() in used:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate.casefold())
    return candidate


def _column_widths(frame: pd.DataFrame, columns: list[str]) -> list[float]:
    sample = frame.loc[:, columns].head(250)
    widths: list[float] = []
    for column in columns:
        values = sample[column].map(lambda value: "" if pd.isna(value) else str(value))
        longest = max([len(str(column)), *values.map(len).tolist()], default=len(str(column)))
        widths.append(float(min(max(longest + 2, 10), 34)))
    return widths


def _number_format(column: str) -> str | None:
    normalized = column.casefold()
    if normalized == "doba_handlowa":
        return "yyyy-mm-dd"
    if any(token in normalized for token in ("timestamp", "_time", "czas", "valid_from", "valid_to")):
        return "yyyy-mm-dd hh:mm:ss"
    if normalized in {
        "wape",
        "smape",
        "udzial",
        "kalibracja_poprawa_mae",
    } or normalized.endswith("_pct"):
        return "0.0%"
    if normalized in {
        "source_row",
        "godzina_handlowa",
        "fold",
        "n",
        "liczba",
        "liczba_foldow",
        "liczba_lagow_bazowych",
        "kalibracja_n",
    }:
        return "#,##0"
    if any(
        token in normalized
        for token in (
            "wartosc",
            "blad",
            "mae",
            "rmse",
            "bias",
            "temperatura",
            "wiatr",
            "zachmurzenie",
            "opad",
            "widocznosc",
            "promieniowanie",
            "albedo",
            "snieg",
            "lead",
            "waznosc",
            "odchylenie",
            "blend_alpha",
        )
    ):
        return "#,##0.00"
    return None


def _append_row(
    worksheet,
    values: Iterable[object],
    columns: list[str],
    body_style,
    date_style,
    integer_style,
    decimal_style,
    percent_style,
) -> None:
    from openpyxl.cell import WriteOnlyCell

    row = []
    for value, column in zip(values, columns):
        cell = WriteOnlyCell(worksheet, value=_excel_scalar(value))
        number_format = _number_format(column)
        if number_format == "yyyy-mm-dd" or number_format == "yyyy-mm-dd hh:mm:ss":
            cell.style = date_style.name
            cell.number_format = number_format
        elif number_format == "#,##0":
            cell.style = integer_style.name
        elif number_format == "0.0%":
            cell.style = percent_style.name
        elif number_format == "#,##0.00":
            cell.style = decimal_style.name
        else:
            cell.style = body_style.name
        row.append(cell)
    worksheet.append(row)


def _write_frame(
    workbook,
    sheet_name: str,
    frame: pd.DataFrame,
    columns: list[str],
    styles: dict[str, object],
) -> None:
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(sheet_name)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    header = []
    for column in columns:
        cell = WriteOnlyCell(worksheet, value=str(column))
        cell.style = styles["header"].name
        header.append(cell)
    worksheet.append(header)

    for values in frame.loc[:, columns].itertuples(index=False, name=None):
        _append_row(
            worksheet,
            values,
            columns,
            styles["body"],
            styles["date"],
            styles["integer"],
            styles["decimal"],
            styles["percent"],
        )

    if columns:
        last_column = get_column_letter(len(columns))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(frame) + 1}"
        for index, width in enumerate(_column_widths(frame, columns), start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width


def _summary_rows(
    predictions: pd.DataFrame,
    metrics: pd.DataFrame,
    models: Iterable[str],
) -> list[list[object]]:
    status = predictions.get("status_predykcji", pd.Series(dtype="string"))
    actual = pd.to_numeric(
        predictions.get("wartosc_rzeczywista", pd.Series(dtype=float)), errors="coerce"
    )
    matched = predictions.get("pogoda_dopasowana", pd.Series(False, index=predictions.index))
    weather_share = float(pd.Series(matched).fillna(False).mean()) if len(predictions) else 0.0
    weather_status = predictions.get(
        "weather_status", pd.Series("BRAK_STATUSU", index=predictions.index)
    ).astype("string")
    post_start_scope = weather_status.isin(["DOPASOWANA", "BRAK_POGODY_W_ZAKRESIE"])
    post_start_weather_share = (
        float(weather_status.loc[post_start_scope].eq("DOPASOWANA").mean())
        if post_start_scope.any()
        else None
    )
    feature_count = pd.to_numeric(
        predictions.get("liczba_cech_pogodowych", pd.Series(0, index=predictions.index)),
        errors="coerce",
    ).fillna(0)
    partial_weather = pd.Series(matched, index=predictions.index).fillna(False) & feature_count.lt(
        len(WEATHER_FEATURES)
    )
    future_without_weather = predictions.get(
        "prognoza_bez_pogody", pd.Series(False, index=predictions.index)
    )

    def metric(model: str, name: str) -> float | None:
        if metrics.empty:
            return None
        selected = metrics[
            metrics["model"].eq(model)
            & metrics["zakres"].eq("GLOBAL")
            & metrics["wartosc_zakresu"].eq("ALL")
        ]
        if selected.empty or name not in selected.columns:
            return None
        value = pd.to_numeric(selected.iloc[0][name], errors="coerce")
        return None if pd.isna(value) else float(value)

    primary_model = (
        "HYBRYDA_OOF"
        if metric("HYBRYDA_OOF", "mae") is not None
        else "ML_OOF"
    )
    hybrid_mae = metric(primary_model, "mae")
    baseline_mae = metric("SREDNIA_ANALOGICZNYCH_D3_D14", "mae")
    improvement = (
        None
        if hybrid_mae is None or baseline_mae in (None, 0)
        else float((baseline_mae - hybrid_mae) / baseline_mae)
    )
    return [
        ["Dane", "Wszystkie wiersze", len(predictions), "Każdy oryginalny wiersz energii"],
        ["Dane", "Wiersze z wykonaniem", int(actual.notna().sum()), "Target Wartość jest dostępny"],
        ["Ocena", "Predykcje OOF hybrydy", int(status.eq("OOF_BACKTEST").sum()), "Tylko te wiersze służą do uczciwej oceny"],
        ["Prognoza", "Wiersze przyszłe", int(status.eq("PROGNOZA_PRZYSZLA").sum()), "Wiersze po ostatnim znanym wykonaniu"],
        ["Prognoza", "Przyszłe bez pogody", int(pd.Series(future_without_weather).fillna(False).sum()), "Predykcja używa historii energii i kalendarza"],
        ["Jakość", "Wiersze przed dostępnością pogody", int(weather_status.eq("PRZED_STARTEM_POGODY").sum()), "Oczekiwany brak przed 2024-10-01"],
        ["Jakość", "Dopasowanie pogody po starcie", post_start_weather_share, "Dopasowane / (dopasowane + brak rekordu); bez błędnych kluczy i mapowania"],
        ["Jakość", "Dopasowanie pogody — wszystkie wiersze", weather_share, "Wskaźnik pomocniczy obejmujący okres sprzed 2024-10-01"],
        ["Jakość", "Dopasowane rekordy z NULL", int(partial_weather.sum()), "Co najmniej jedna z cech pogodowych jest pusta; wiersz pozostaje"],
        ["Hybryda", "Wytrenowane kierunki", ", ".join(sorted(models)) or "brak", "Oddzielne korekty residualne pobrania i oddania"],
        ["Hybryda", "MAE hybrydy (OOF)", hybrid_mae, "Niżej = lepiej; jednostka taka jak Wartość"],
        ["Baseline", "MAE średniej D-3...D-14", baseline_mae, "Średnia analogicznej godziny z 3–14 dni wcześniej"],
        ["Hybryda", "Poprawa MAE vs baseline", improvement, "Korekta ML jest używana tylko po spełnieniu progu poprawy"],
        ["Hybryda", "WAPE hybrydy (OOF)", metric(primary_model, "wape"), "Błąd bezwzględny względem wolumenu"],
        ["Baseline", "WAPE średniej D-3...D-14", metric("SREDNIA_ANALOGICZNYCH_D3_D14", "wape"), "Punkt odniesienia"],
    ]


def write_results_workbook(
    predictions: pd.DataFrame,
    prediction_columns: list[str],
    metrics: pd.DataFrame,
    feature_importance: pd.DataFrame,
    quality: pd.DataFrame,
    mapping: pd.DataFrame,
    models: Iterable[str],
    manifest: dict[str, object],
    output_path: str | Path,
    summary_predictions: pd.DataFrame | None = None,
) -> Path:
    """Zapisuje jeden audytowalny XLSX; duże predykcje dzieli na arkusze."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Do zapisu wyniki_mdd.xlsx wymagany jest openpyxl.") from exc

    workbook = Workbook(write_only=True)
    thin_gray = Side(style="thin", color="D9E2F3")

    header = NamedStyle(name="mdd_header")
    header.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="1F4E78")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.border = Border(bottom=Side(style="medium", color="17365D"))

    body = NamedStyle(name="mdd_body")
    body.font = Font(name="Aptos", size=10, color="1F2937")
    body.alignment = Alignment(vertical="top")
    body.border = Border(bottom=thin_gray)

    date_style = NamedStyle(name="mdd_date")
    date_style.font = body.font
    date_style.alignment = Alignment(horizontal="center", vertical="top")
    date_style.border = body.border

    integer = NamedStyle(name="mdd_integer")
    integer.font = body.font
    integer.alignment = Alignment(horizontal="right", vertical="top")
    integer.border = body.border
    integer.number_format = "#,##0"

    decimal = NamedStyle(name="mdd_decimal")
    decimal.font = body.font
    decimal.alignment = Alignment(horizontal="right", vertical="top")
    decimal.border = body.border
    decimal.number_format = "#,##0.00"

    percent = NamedStyle(name="mdd_percent")
    percent.font = body.font
    percent.alignment = Alignment(horizontal="right", vertical="top")
    percent.border = body.border
    percent.number_format = "0.0%"

    for style in (header, body, date_style, integer, decimal, percent):
        workbook.add_named_style(style)
    styles = {
        "header": header,
        "body": body,
        "date": date_style,
        "integer": integer,
        "decimal": decimal,
        "percent": percent,
    }

    used_names: set[str] = set()
    summary_source = (
        predictions if summary_predictions is None else summary_predictions
    )
    summary = pd.DataFrame(
        _summary_rows(summary_source, metrics, models),
        columns=["sekcja", "wskaznik", "wartosc", "opis"],
    )
    _write_frame(
        workbook,
        _safe_sheet_name("Podsumowanie", used_names),
        summary,
        list(summary.columns),
        styles,
    )

    prediction_view = predictions.loc[:, prediction_columns]
    if "source_sheet" in prediction_view.columns:
        groups = prediction_view.groupby("source_sheet", dropna=False, sort=False)
    else:
        groups = [("Predykcje", prediction_view)]
    for source_sheet, group in groups:
        for chunk_no, start in enumerate(range(0, len(group), EXCEL_DATA_ROWS_PER_SHEET), 1):
            chunk = group.iloc[start : start + EXCEL_DATA_ROWS_PER_SHEET]
            suffix = "" if chunk_no == 1 else f"_{chunk_no}"
            name = _safe_sheet_name(f"Pred_{source_sheet}{suffix}", used_names)
            _write_frame(workbook, name, chunk, prediction_columns, styles)

    for name, frame in [
        ("Metryki", metrics),
        ("Waznosc_cech", feature_importance),
        ("Kontrola_jakosci", quality),
        ("Mapowanie", mapping),
    ]:
        _write_frame(
            workbook,
            _safe_sheet_name(name, used_names),
            frame,
            list(frame.columns),
            styles,
        )

    manifest_rows = [["klucz", "wartosc"]]
    for key, value in manifest.items():
        manifest_rows.append(
            [str(key), json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)]
        )
    manifest_frame = pd.DataFrame(manifest_rows[1:], columns=manifest_rows[0])
    _write_frame(
        workbook,
        _safe_sheet_name("Konfiguracja", used_names),
        manifest_frame,
        list(manifest_frame.columns),
        styles,
    )

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
